import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import os
import yaml
import time
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL

# ==========================================
# --- PAGE CONFIGURATION & DESIGN SYSTEM ---
# ==========================================
st.set_page_config(page_title="OpsFlow Studio", page_icon="⚡", layout="wide")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }
    
    .stDownloadButton button, .stButton button {
        width: 100%;
        border-radius: 6px !important;
        font-weight: 500 !important;
        transition: all 0.15s ease-in-out !important;
    }
    
    .stButton button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06) !important;
    }
    
    [data-testid="stExpander"] {
        border-radius: 8px !important;
        border: 1px solid #e5e7eb !important;
        box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05) !important;
    }
    
    .block-container {
        padding-top: 2rem !important;
    }
    header {visibility: hidden;}
    footer {visibility: hidden;}
    
    h1 { font-weight: 600 !important; letter-spacing: -0.025em !important; }
    h2, h3 { font-weight: 500 !important; letter-spacing: -0.025em !important; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# --- CONFIGURATION & DATABASE ---
# ==========================================
try:
    with open("config.yaml", "r") as file:
        config = yaml.safe_load(file)
    db_cfg = config['database']
except FileNotFoundError:
    st.error("Configuration file 'config.yaml' not found. Please create it.")
    st.stop()

url_object = URL.create(
    drivername="postgresql+psycopg2",
    username=db_cfg['username'],
    password=db_cfg['password'],
    host=db_cfg['host'],
    port=db_cfg['port'],
    database=db_cfg['dbname']
)

engine = create_engine(url_object)

def init_db():
    try:
        with engine.begin() as conn:
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS run_history (
                    id SERIAL PRIMARY KEY,
                    report_type VARCHAR(50),
                    run_date VARCHAR(50),
                    filename VARCHAR(255),
                    record_count INT,
                    file_data BYTEA
                )
            '''))
    except Exception as e:
        print(f"Database Initialization Error: {e}")

def save_run_to_db(report_type, filename, record_count, file_data):
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO run_history (report_type, run_date, filename, record_count, file_data) 
                VALUES (:report_type, :run_date, :filename, :record_count, :file_data)
            """),
            {
                "report_type": report_type, 
                "run_date": run_date, 
                "filename": filename, 
                "record_count": record_count, 
                "file_data": file_data
            }
        )

def get_history_metadata(report_type):
    try:
        query = text("SELECT id, run_date, filename, record_count FROM run_history WHERE report_type = :rtype ORDER BY run_date DESC")
        df = pd.read_sql(query, engine, params={"rtype": report_type})
        return df
    except Exception as e:
        return pd.DataFrame(columns=['id', 'run_date', 'filename', 'record_count'])

def get_file_from_db(record_id):
    try:
        query = text("SELECT file_data, filename FROM run_history WHERE id = :rid")
        with engine.connect() as conn:
            result = conn.execute(query, {"rid": record_id}).fetchone()
            if result:
                return bytes(result[0]), result[1]
    except Exception as e:
        pass
    return None, None

def delete_record_from_db(record_id):
    try:
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM run_history WHERE id = :rid"), {"rid": record_id})
        return True
    except Exception as e:
        return False

def export_to_local_output(data, filename):
    os.makedirs("output", exist_ok=True)
    filepath = os.path.join("output", filename)
    with open(filepath, "wb") as f:
        if isinstance(data, io.BytesIO):
            f.write(data.getvalue())
        else:
            f.write(data)
    return filepath

init_db()

# ==========================================
# --- STATE INITIALIZATION ---
# ==========================================
if "doms_counter" not in st.session_state: st.session_state.doms_counter = 0
if "myf_counter" not in st.session_state: st.session_state.myf_counter = 0
if "inc_counter" not in st.session_state: st.session_state.inc_counter = 0
if "scanner_counter" not in st.session_state: st.session_state.scanner_counter = 0

# Persistent state for the Scanner Editor to prevent data loss on filter changes
if "scanner_master_df" not in st.session_state: st.session_state.scanner_master_df = None
if "scanner_audit_df" not in st.session_state: st.session_state.scanner_audit_df = None
if "scanner_file_id" not in st.session_state: st.session_state.scanner_file_id = ""

# ==========================================
# --- CORE PROCESSING LOGIC ---
# ==========================================
def extract_first_number(area_str):
    if not isinstance(area_str, str):
        return None
    match = re.search(r'\d+', area_str)
    return int(match.group()) if match else None

def process_files(file1, file2):
    if file1.name.endswith('.csv'):
        df1 = pd.read_csv(file1)
    else:
        df1 = pd.read_excel(file1)
        
    df1.rename(columns={'Resolution Notes': 'Resolution Note'}, inplace=True)
    
    for col in ['Description', 'Cause Category', 'Cause Code']:
        if col not in df1.columns:
            df1[col] = ''
    
    xls = pd.ExcelFile(file2)
    df2 = pd.read_excel(file2, sheet_name=xls.sheet_names[0])
    df2.columns = df2.columns.str.strip()
    df2['Station No.'] = pd.to_numeric(df2['Station No.'], errors='coerce').astype('Int64')
    
    df1['Extracted_Station_No'] = df1['Area'].apply(extract_first_number).astype('Int64')
    
    merged_df = pd.merge(df1, df2, left_on='Extracted_Station_No', right_on='Station No.', how='inner')
    
    merged_df['Created_dt'] = pd.to_datetime(merged_df['Created'], errors='coerce')
    merged_df['Date_dt'] = pd.to_datetime(merged_df['Date'], errors='coerce')
    filtered_df = merged_df[merged_df['Created_dt'] >= merged_df['Date_dt']]
    
    res_note = filtered_df['Resolution Note'].fillna('').str.lower()
    summary = filtered_df['Summary'].fillna('').str.lower()
    desc = filtered_df['Description'].fillna('').str.lower()
    
    has_rfid = res_note.str.contains('rfid') | summary.str.contains('rfid') | desc.str.contains('rfid')
    has_doms_pump = res_note.str.contains('doms|pump') | summary.str.contains('doms|pump') | desc.str.contains('doms|pump')
    has_offline = res_note.str.contains(r'pss?\s*5000\s*offline\s*\(\s*1100\s*\)') | \
                  summary.str.contains(r'pss?\s*5000\s*offline\s*\(\s*1100\s*\)') | \
                  desc.str.contains(r'pss?\s*5000\s*offline\s*\(\s*1100\s*\)')
    
    target_mask = has_doms_pump & ~has_rfid & ~has_offline
    filtered_df = filtered_df[target_mask]
    
    final_cols = [
        'Station No.', 'Number', 'Priority', 'Created', 'Summary', 
        'Resolution Note', 'Status', 'Cause Category', 'Cause Code', 
        'DOMS Model', 'Date'
    ]
    
    final_cols = [col for col in final_cols if col in filtered_df.columns]
    final_df = filtered_df[final_cols]
    
    return final_df

def process_scanner_comparison(tracker_file, scan_file):
    if tracker_file.name.endswith('.csv'):
        df_tracker = pd.read_csv(tracker_file)
    else:
        df_tracker = pd.read_excel(tracker_file)
        
    if scan_file.name.endswith('.csv'):
        df_scan = pd.read_csv(scan_file)
    else:
        df_scan = pd.read_excel(scan_file)
        
    df_tracker.columns = df_tracker.columns.astype(str).str.strip()
    df_scan.columns = df_scan.columns.astype(str).str.strip()
    
    # 1. FILTER: Remove any rows containing "Connection Failed" BEFORE manipulating columns
    df_tracker = df_tracker[~df_tracker.apply(lambda col: col.astype(str).str.contains('Connection Failed', case=False, na=False)).any(axis=1)]
    df_scan = df_scan[~df_scan.apply(lambda col: col.astype(str).str.contains('Connection Failed', case=False, na=False)).any(axis=1)]

    # 1.5. FUZZY COLUMN MATCHER: Synchronizes names like "POS Manufacture" and "POS Manufacturer"
    def normalize_name(name):
        return str(name).strip().lower().replace(" ", "").replace("\n", "").replace("manufacturer", "manufacture")

    scan_col_map = {normalize_name(col): col for col in df_scan.columns}
    tracker_rename_map = {}
    
    for col in df_tracker.columns:
        norm_col = normalize_name(col)
        if norm_col in scan_col_map:
            tracker_rename_map[col] = scan_col_map[norm_col]
            
    # Rename tracker columns so they perfectly match the scan columns
    df_tracker.rename(columns=tracker_rename_map, inplace=True)

    # 2. Ignore Unnecessary Operational Columns entirely
    ignore_cols = ['Engineer', 'status', 'Status', 'Note', 'Notes']
    df_tracker = df_tracker.drop(columns=[c for c in ignore_cols if c in df_tracker.columns], errors='ignore')
    df_scan = df_scan.drop(columns=[c for c in ignore_cols if c in df_scan.columns], errors='ignore')
    
    join_key = 'POS Name'
    # Fallback to fuzzy search for join key if exact match fails
    if join_key not in df_scan.columns:
        for col in df_scan.columns:
            if normalize_name(col) == normalize_name(join_key):
                join_key = col
                break

    if join_key not in df_tracker.columns or join_key not in df_scan.columns:
        raise ValueError(f"Required identifier column '{join_key}' not found in both files.")

    # PRE-MERGE FIX: Drop duplicate rows based on POS Name to prevent Cartesian products
    df_scan = df_scan.drop_duplicates(subset=[join_key], keep='first')
    df_tracker = df_tracker.drop_duplicates(subset=[join_key], keep='first')

    merged_df = pd.merge(df_scan, df_tracker, on=join_key, how='outer', suffixes=('_scan', '_tracker'))
    audit_records = []
    
    # Define synonyms for missing values, explicitly handling "not configured"
    null_synonyms = ['unknown', 'nan', 'nat', 'none', 'null', '', 'not configured']
    
    # 3. Reconcile "Unknown" & "NOT CONFIGURED" fields 
    for col in df_scan.columns:
        if col != join_key and f"{col}_tracker" in merged_df.columns:
            scan_col = f"{col}_scan"
            tracker_col = f"{col}_tracker"
            
            # Stringify and thoroughly clean to catch all variations of Empty/None/NaN/NOT CONFIGURED
            scan_str = merged_df[scan_col].astype(str).str.strip().str.lower()
            tracker_str = merged_df[tracker_col].astype(str).str.strip().str.lower()
            
            is_unknown_in_scan = scan_str.isin(null_synonyms) | merged_df[scan_col].isna()
            
            # Check if Tracker has a valid fallback value
            has_value_in_tracker = merged_df[tracker_col].notna() & ~tracker_str.isin(null_synonyms)
            
            # EXCLUSION RULE: Exclude patching DR9401730 specifically for the ApplicationROMID column
            if 'applicationromid' in normalize_name(col):
                is_excluded = merged_df[tracker_col].astype(str).str.strip().str.upper() == 'DR9401730'
                has_value_in_tracker = has_value_in_tracker & ~is_excluded
            
            # Identify rows to patch
            change_mask = is_unknown_in_scan & has_value_in_tracker
            
            # Log changes to the audit trail (will only display in the app)
            if change_mask.any():
                changed_rows = merged_df[change_mask]
                for idx, row in changed_rows.iterrows():
                    audit_records.append({
                        'POS Name': row[join_key],
                        'Field Name': col,
                        'Before (Scan)': str(row[scan_col]),
                        'After (Tracker)': str(row[tracker_col])
                    })
            
            # Convert both columns to a generic 'object' type to prevent DateTime / Float crash
            merged_df[col] = np.where(
                change_mask, 
                merged_df[tracker_col].astype(object), 
                merged_df[scan_col].astype(object)
            )
            merged_df.drop(columns=[scan_col, tracker_col], inplace=True, errors='ignore')

    # 4. Restrict Export Output, Add POS Model, Remove Interface, and Enforce Order
    target_columns_clean = [
        'posname',
        'posmanufacture',
        'posmodel',
        'latestoposinstalled',
        'currentscannerdriverinnamos',
        'applicationromid(fw)',
        'configurationfileid'
    ]
    
    final_cols = []
    seen_columns = set()
    
    # Loop through merged columns, matching them strictly to the clean target list
    for original_col in merged_df.columns:
        clean_col = normalize_name(original_col)
        if clean_col in target_columns_clean and clean_col not in seen_columns:
            final_cols.append(original_col)
            seen_columns.add(clean_col)
            
    merged_df = merged_df[final_cols]
    
    # Standardize column names to the exact required presentation names
    final_rename_map = {}
    for col in merged_df.columns:
        norm = normalize_name(col)
        if norm == 'posname': final_rename_map[col] = 'POS Name'
        elif norm == 'posmanufacture': final_rename_map[col] = 'POS Manufacture'
        elif norm == 'posmodel': final_rename_map[col] = 'POS Model'
        elif norm == 'latestoposinstalled': final_rename_map[col] = 'Latest OPOS Installed'
        elif norm == 'currentscannerdriverinnamos': final_rename_map[col] = 'Current Scanner Driver in Namos'
        elif 'applicationromid' in norm: final_rename_map[col] = 'ApplicationROMID (FW)'
        elif norm == 'configurationfileid': final_rename_map[col] = 'ConfigurationFileID'
        
    merged_df.rename(columns=final_rename_map, inplace=True)
    
    # Enforce strict column order for the final output
    final_order = [
        'POS Name',
        'POS Manufacture',
        'POS Model',
        'Latest OPOS Installed',
        'Current Scanner Driver in Namos',
        'ApplicationROMID (FW)',
        'ConfigurationFileID'
    ]
    available_cols = [c for c in final_order if c in merged_df.columns]
    merged_df = merged_df[available_cols]
    
    # Double-safety net to drop any remaining duplicate columns created by pandas merge suffixes
    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
    merged_df = merged_df.drop_duplicates(subset=['POS Name'], keep='first')
    
    # Reset index strictly so mapping edits back across states is robust
    merged_df.reset_index(drop=True, inplace=True)
    merged_df = merged_df.fillna("")
            
    df_audit = pd.DataFrame(audit_records)
    return merged_df, df_audit

def generate_excel_bytes(df):
    output = io.BytesIO()
    wb = Workbook()
    
    # --- Sheet 1: Main Reconciled Results ---
    ws1 = wb.active
    ws1.title = "Reconciled Devices"
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
        
    # --- Sheet 2: Pivot Table Summary (With Station Count) ---
    ws2 = wb.create_sheet(title="Pivot Summary")
    
    pivot_req_cols = ['POS Manufacture', 'ApplicationROMID (FW)', 'ConfigurationFileID']
    if all(col in df.columns for col in pivot_req_cols) and 'POS Name' in df.columns:
        pivot_data = df.copy()
        
        # Extract Station ID by stripping out "-POS" and trailing numbers
        pivot_data['Station ID'] = pivot_data['POS Name'].apply(
            lambda x: re.split(r'-POS\d*', str(x), flags=re.IGNORECASE)[0] if pd.notna(x) else 'Unknown'
        )
        
        pivot_data[pivot_req_cols] = pivot_data[pivot_req_cols].replace("", "Unknown").fillna("Unknown")
        
        # Group by the pivot columns and calculate both POS counts and unique Station counts
        pivot_df = pivot_data.groupby(pivot_req_cols).agg(
            Device_Count=('POS Name', 'count'),
            Station_Count=('Station ID', 'nunique')
        ).reset_index()
        
        # Rename columns for the final display
        pivot_df.rename(columns={
            'Device_Count': 'Total POS Devices', 
            'Station_Count': 'Total Unique Stations'
        }, inplace=True)
        
        # Sort values nicely to replicate standard Excel hierarchical pivot order
        pivot_df = pivot_df.sort_values(by=['POS Manufacture', 'ApplicationROMID (FW)'])
        
        for r in dataframe_to_rows(pivot_df, index=False, header=True):
            ws2.append(r)
    else:
        ws2.append(["Required columns missing for Pivot generation."])

    # --- Formatting Variables ---
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin', color='E5E7EB'),
                         right=Side(style='thin', color='E5E7EB'),
                         top=Side(style='thin', color='E5E7EB'),
                         bottom=Side(style='thin', color='E5E7EB'))
                         
    # --- Apply uniform formatting to both worksheets ---
    for ws in [ws1, ws2]:
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
        for col in ws.columns:
            max_length = max((len(str(cell.value).split('\n')[0]) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(min(max_length + 2, 65), 18)
            
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# --- UI: SIDEBAR & NAVIGATION ---
# ==========================================
st.title("⚡ OpsFlow Studio")
st.markdown("Automated ticket reconciliation, audit pipelines, and scanner telemetry management.")

with st.sidebar:
    st.header("Workspace Guide")
    st.markdown("---")
    st.markdown("""
    **DOMS Reconciliation**
    *   Isolates tickets mentioning `DOMS` or `Pumps`.
    *   Excludes `RFID` references.
    
    **MYF & Incidents**
    *   Cross-references raw incidents against audited MYF datasets.
    
    **Scanner Recon & Patching**
    *   Compares fresh scan data against tracker sheets.
    *   Auto-fills `Unknown`, blanks, and `NOT CONFIGURED` values using tracker history.
    *   **Persistent Editor:** Safely retains edits even when switching filters.
    *   **Exclusions:** Will not patch `ApplicationROMID (FW)` if the tracker value is `DR9401730`.
    *   **Data Export:** Drops `Connection Failed`, strictly orders 7 core columns, and exports an automated Pivot Summary with POS and Station counts.
    """)
    st.markdown("---")
    st.caption("Engine: PostgreSQL Desktop | Docker")

tab1, tab2, tab3 = st.tabs(["DOMS Data Merge", "MYF & Incident Management", "Scanner Recon & Patching"])

# ==========================================
# --- TAB 1: DOMS RECONCILIATION ---
# ==========================================
with tab1:
    st.markdown("<br>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("1. Source Data")
        col_input1, col_input2 = st.columns(2)
        with col_input1:
            file1 = st.file_uploader("Raw Incidents Report (CSV/XLSX)", type=["csv", "xlsx"], key=f"doms_f1_{st.session_state.doms_counter}")
        with col_input2:
            file2 = st.file_uploader("DOMS Rollout Schedule (XLSX)", type=["xlsx"], key=f"doms_f2_{st.session_state.doms_counter}")

    if file1 and file2:
        with st.spinner('Reconciling datasets...'):
            try:
                raw_final_df = process_files(file1, file2)
                row_count_initial = len(raw_final_df)
                
                if row_count_initial == 0:
                    st.warning("⚠️ No matching DOMS/PUMPS records found based on exclusion rules.")
                else:
                    st.toast(f"✅ Reconciled {row_count_initial} valid records.", icon="✅")
                    with st.container(border=True):
                        st.subheader("2. Audit & Verify")
                        raw_final_df.insert(0, '🗑️ Remove', False)
                        edited_preview_df = st.data_editor(raw_final_df, num_rows="fixed", width='stretch', hide_index=True, key="doms_editor")
                        
                        clean_preview_df = edited_preview_df[~edited_preview_df['🗑️ Remove']].drop(columns=['🗑️ Remove'])
                        current_row_count = len(clean_preview_df)
                        excel_data = generate_excel_bytes(clean_preview_df)
                        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                        export_filename = f"DOMS_Recon_{timestamp_str}.xlsx"
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        col_act1, col_act2 = st.columns([3, 1])
                        with col_act1:
                            if st.button("💾 Commit & Save to Database", key="save_doms_db", type="primary"):
                                save_run_to_db("DOMS", export_filename, current_row_count, excel_data.getvalue())
                                st.toast("Commit successful.", icon="✅")
                                time.sleep(1)
                                st.session_state.doms_counter += 1
                                st.rerun()
                        with col_act2:
                            if st.button("📥 Direct Export", key="doms_dl_new"):
                                filepath = export_to_local_output(excel_data, export_filename)
                                st.toast(f"Exported to /output directory.", icon="🚀")
                                time.sleep(1)
                                st.session_state.doms_counter += 1
                                st.rerun()
            except Exception as e:
                st.error(f"Execution Error: {e}")
                
    st.markdown("<br>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("🗄️ Extraction History")
        history_df = get_history_metadata("DOMS")
        
        if history_df.empty:
            st.caption("No historical deployments found.")
        else:
            for index, row in history_df.iterrows():
                r_id = row['id']
                r_col1, r_col2, r_col3, r_col4, r_col5, r_col6 = st.columns([2, 3, 1, 1, 1, 1], vertical_alignment="center")
                with r_col1: st.caption(str(row['run_date']))
                with r_col2: st.caption(str(row['filename']))
                with r_col3: st.caption(f"{row['record_count']}")
                
                db_file_data, db_filename = get_file_from_db(r_id)
                with r_col4:
                    if db_file_data and st.button("📥 Export", key=f"dl_doms_{r_id}", use_container_width=True):
                        export_to_local_output(db_file_data, db_filename)
                        st.toast("Exported to local volume.", icon="🚀")
                        time.sleep(1)
                        st.rerun()
                with r_col5:
                    if st.button("👁️ View", key=f"rev_doms_{r_id}", use_container_width=True):
                        st.session_state[f"show_review_{r_id}"] = not st.session_state.get(f"show_review_{r_id}", False)
                with r_col6:
                    if st.button("🗑️ Drop", key=f"del_doms_{r_id}", use_container_width=True):
                        delete_record_from_db(r_id)
                        st.rerun()
                
                if st.session_state.get(f"show_review_{r_id}", False) and db_file_data:
                    st.info(f"Viewing Snapshot: {db_filename}")
                    st.dataframe(pd.read_excel(io.BytesIO(db_file_data)), width='stretch')
                st.divider()

# ==========================================
# --- TAB 2: MYF & INCIDENT MANAGEMENT ---
# ==========================================
with tab2:
    st.markdown("<br>", unsafe_allow_html=True)
    subtab1, subtab2 = st.tabs(["Step 1: MYF Context Editor", "Step 2: Incident Cross-Match"])
    
    with subtab1:
        with st.container(border=True):
            st.subheader("MYF Initialization")
            myf_file = st.file_uploader("Upload Generated MYF Extraction (Excel)", type=["xlsx"], key=f"myf_upload_{st.session_state.myf_counter}")
            
            if 'myf_reference_df' not in st.session_state:
                st.session_state.myf_reference_df = None

            if myf_file:
                try:
                    myf_bytes = io.BytesIO(myf_file.read())
                    xls_myf = pd.ExcelFile(myf_bytes)
                    sheet_names = xls_myf.sheet_names
                    
                    selected_sheet = st.selectbox("Target Worksheet", sheet_names, key="myf_sheet_select")
                    df_myf = pd.read_excel(myf_bytes, sheet_name=selected_sheet)
                    st.session_state.myf_reference_df = df_myf.copy()
                    
                    df_myf.insert(0, '🗑️ Remove', False)
                    date_cols = ['Received Time', 'Action Start Time', 'Action End Time']
                    for col in date_cols:
                        if col in df_myf.columns:
                            df_myf[col] = pd.to_datetime(df_myf[col], format='mixed', errors='coerce')
                    
                    column_configurations = {'🗑️ Remove': st.column_config.CheckboxColumn("🗑️ Remove", default=False)}
                    for col in date_cols:
                        if col in df_myf.columns:
                            column_configurations[col] = st.column_config.DatetimeColumn(col, format="DD-MM-YYYY HH:mm:ss", step=1)
                    
                    edited_myf_df = st.data_editor(df_myf, num_rows="fixed", width='stretch', hide_index=True, column_config=column_configurations, key="myf_editor")
                    clean_myf_df = edited_myf_df[~edited_myf_df['🗑️ Remove']].drop(columns=['🗑️ Remove'])
                    
                    myf_output = io.BytesIO()
                    with pd.ExcelWriter(myf_output, engine='openpyxl') as writer:
                        clean_myf_df.to_excel(writer, index=False, sheet_name='Engineers Updated')
                    
                    myf_filename = f"MYF_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    col_m1, col_m2 = st.columns([3, 1])
                    with col_m1:
                        if st.button("💾 Commit MYF to Database", key="save_myf_db", type="primary"):
                            save_run_to_db("MYF_Update", myf_filename, len(clean_myf_df), myf_output.getvalue())
                            st.toast("MYF context saved.", icon="✅")
                            time.sleep(1)
                            st.session_state.myf_counter += 1
                            st.rerun()
                    with col_m2:
                        if st.button("📥 Direct Export", key="myf_download_btn"):
                            export_to_local_output(myf_output, myf_filename)
                            st.toast("Exported to /output.", icon="🚀")
                            time.sleep(1)
                            st.session_state.myf_counter += 1
                            st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

    with subtab2:
        with st.container(border=True):
            st.subheader("Incident Cross-Reference Engine")
            if st.session_state.myf_reference_df is None:
                st.info("🔒 Prerequisite: Upload MYF Results in Step 1 first.")
            else:
                incident_file = st.file_uploader("Raw Incident Report (CSV/XLSX)", type=["csv", "xlsx"], key=f"inc_upload_{st.session_state.inc_counter}")
                if incident_file:
                    try:
                        df_raw_inc = pd.read_csv(incident_file) if incident_file.name.endswith('.csv') else pd.read_excel(incident_file)
                        df_raw_inc.rename(columns={'Resolution Note': 'Resolution Notes'}, inplace=True)
                        
                        myf_df = st.session_state.myf_reference_df
                        valid_tickets = set(myf_df['Incident Number'].dropna().astype(str).str.strip().str.upper()) if 'Incident Number' in myf_df.columns else set()
                        
                        inc_number_col = next((c for c in ['Number', 'Incident Number', 'Incident'] if c in df_raw_inc.columns), None)
                        if inc_number_col:
                            df_matched = df_raw_inc[df_raw_inc[inc_number_col].astype(str).str.strip().str.upper().isin(valid_tickets)]
                            df_matched.insert(0, '🗑️ Remove', False)
                            edited_matched_inc = st.data_editor(df_matched, num_rows="fixed", width='stretch', hide_index=True, key="matched_inc_editor")
                            
                            clean_matched_inc = edited_matched_inc[~edited_matched_inc['🗑️ Remove']].drop(columns=['🗑️ Remove'])
                            inc_output = io.BytesIO()
                            with pd.ExcelWriter(inc_output, engine='openpyxl') as writer:
                                clean_matched_inc.to_excel(writer, index=False, sheet_name='Formatted_Incidents')
                            
                            inc_filename = f"Matched_Incidents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                            
                            if st.button("💾 Commit Matched Report", key="save_matched_db", type="primary"):
                                save_run_to_db("Matched_Incident", inc_filename, len(clean_matched_inc), inc_output.getvalue())
                                st.toast("Saved successfully.", icon="✅")
                                time.sleep(1)
                                st.session_state.inc_counter += 1
                                st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

    st.markdown("<br>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("🗄️ MYF & Match History")
        history_combined = pd.concat([get_history_metadata("MYF_Update"), get_history_metadata("Matched_Incident")])
        if not history_combined.empty:
            history_combined = history_combined.sort_values(by="run_date", ascending=False)
            for index, row in history_combined.iterrows():
                r_id = row['id']
                r_col1, r_col2, r_col3, r_col4, r_col5, r_col6 = st.columns([2, 3, 1, 1, 1, 1], vertical_alignment="center")
                with r_col1: st.caption(str(row['run_date']))
                with r_col2: st.caption(str(row['filename']))
                with r_col3: st.caption(f"{row['record_count']}")
                db_file_data, db_filename = get_file_from_db(r_id)
                with r_col4:
                    if db_file_data and st.button("📥 Export", key=f"dl_tab2_{r_id}", use_container_width=True):
                        export_to_local_output(db_file_data, db_filename)
                        st.toast("Exported.", icon="🚀")
                        time.sleep(1)
                        st.rerun()
                with r_col5:
                    if st.button("👁️ View", key=f"rev_tab2_{r_id}", use_container_width=True):
                        st.session_state[f"show_review_{r_id}"] = not st.session_state.get(f"show_review_{r_id}", False)
                with r_col6:
                    if st.button("🗑️ Drop", key=f"del_tab2_{r_id}", use_container_width=True):
                        delete_record_from_db(r_id)
                        st.rerun()
                if st.session_state.get(f"show_review_{r_id}", False) and db_file_data:
                    st.dataframe(pd.read_excel(io.BytesIO(db_file_data)), width='stretch')
                st.divider()

# ==========================================
# --- TAB 3: SCANNER RECON & PATCHING ---
# ==========================================
with tab3:
    st.markdown("<br>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("1. Scanner Telemetry Ingestion")
        st.markdown("Compare your updated tracker sheet against a fresh device scan. Fields showing `Unknown`, blank, or `NOT CONFIGURED` in the fresh scan will be automatically patched using the tracker data.")
        
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            tracker_file = st.file_uploader("Updated Tracker Sheet (XLSX/CSV)", type=["csv", "xlsx"], key=f"tracker_file_{st.session_state.scanner_counter}")
        with col_s2:
            scan_file = st.file_uploader("Fresh Device Scan Sheet (XLSX/CSV)", type=["csv", "xlsx"], key=f"scan_file_{st.session_state.scanner_counter}")

    if tracker_file and scan_file:
        with st.spinner('Reconciling and patching telemetry data...'):
            try:
                # Create a unique ID for this upload pair to process once
                current_file_id = f"{tracker_file.name}_{scan_file.name}_{st.session_state.scanner_counter}"
                
                # Check if this is a fresh upload. If so, process and save to master state
                if st.session_state.scanner_file_id != current_file_id:
                    reconciled_df, audit_df = process_scanner_comparison(tracker_file, scan_file)
                    reconciled_df.insert(0, '🗑️ Remove', False)
                    
                    st.session_state.scanner_master_df = reconciled_df
                    st.session_state.scanner_audit_df = audit_df
                    st.session_state.scanner_file_id = current_file_id

                if st.session_state.scanner_master_df is not None:
                    st.success(f"✅ Reconciliation Complete! Successfully processed {len(st.session_state.scanner_master_df)} devices.")
                    
                    if not st.session_state.scanner_audit_df.empty:
                        with st.expander("🔍 View Before & After Audit Trail (App-Only)", expanded=False):
                            st.caption("The following fields were patched because the fresh scan reported 'Unknown', blank, or 'NOT CONFIGURED', but valid tracker history existed:")
                            st.dataframe(st.session_state.scanner_audit_df, width='stretch', hide_index=True)
                    else:
                        st.info("ℹ️ No missing values found in the fresh scan that required patching from the tracker.")

                    with st.container(border=True):
                        st.subheader("2. Reconciled Preview & Final Editor")
                        
                        # --- SMART FILTER PANEL (EXCEL STYLE) ---
                        st.markdown("##### 🔍 Filter Records")
                        f_col1, f_col2, f_col3, f_col4 = st.columns(4)
                        
                        master_df = st.session_state.scanner_master_df
                        
                        with f_col1:
                            mfg_options = sorted([str(x) for x in master_df['POS Manufacture'].unique() if pd.notna(x) and str(x).strip() != ''])
                            mfg_filter = st.multiselect("POS Manufacture", options=mfg_options)
                            
                        with f_col2:
                            model_options = sorted([str(x) for x in master_df['POS Model'].unique() if pd.notna(x) and str(x).strip() != ''])
                            model_filter = st.multiselect("POS Model", options=model_options)
                            
                        with f_col3:
                            fw_options = sorted([str(x) for x in master_df['ApplicationROMID (FW)'].unique() if pd.notna(x) and str(x).strip() != ''])
                            fw_filter = st.multiselect("Firmware (ROMID)", options=fw_options)
                            
                        with f_col4:
                            search_query = st.text_input("Search POS Name")
                            
                        # Apply Filters securely while preserving the master index mapping
                        filtered_df = master_df.copy()
                        
                        if mfg_filter:
                            filtered_df = filtered_df[filtered_df['POS Manufacture'].isin(mfg_filter)]
                        if model_filter:
                            filtered_df = filtered_df[filtered_df['POS Model'].isin(model_filter)]
                        if fw_filter:
                            filtered_df = filtered_df[filtered_df['ApplicationROMID (FW)'].isin(fw_filter)]
                        if search_query:
                            filtered_df = filtered_df[filtered_df['POS Name'].str.contains(search_query, case=False, na=False)]
                            
                        st.caption(f"Showing {len(filtered_df)} of {len(master_df)} total records.")

                        # Pass ONLY the filtered view to the editor
                        edited_filtered_df = st.data_editor(filtered_df, num_rows="fixed", width='stretch', hide_index=True, key=f"scanner_editor_{st.session_state.scanner_counter}")
                        
                        # --- MERGE EDITS BACK TO FULL DATASET IN SESSION STATE ---
                        if not edited_filtered_df.empty:
                            # Map changes perfectly back over their original rows in the master dataframe
                            st.session_state.scanner_master_df.loc[edited_filtered_df.index, edited_filtered_df.columns] = edited_filtered_df
                        
                        # Final cleaned dataset used for saving and exporting
                        clean_scanner_df = st.session_state.scanner_master_df[~st.session_state.scanner_master_df['🗑️ Remove']].drop(columns=['🗑️ Remove'])
                        scanner_row_count = len(clean_scanner_df)
                        
                        # --- LIVE UI PIVOT SUMMARY ---
                        st.markdown("<hr>", unsafe_allow_html=True)
                        st.subheader("📊 Live Pivot Summary")
                        st.caption("Instantly recalculates as you modify the data above. This summary will be included as 'Sheet 2' in your Excel export.")
                        
                        pivot_req_cols = ['POS Manufacture', 'ApplicationROMID (FW)', 'ConfigurationFileID']
                        if all(col in clean_scanner_df.columns for col in pivot_req_cols) and 'POS Name' in clean_scanner_df.columns:
                            live_pivot_data = clean_scanner_df.copy()
                            live_pivot_data['Station ID'] = live_pivot_data['POS Name'].apply(
                                lambda x: re.split(r'-POS\d*', str(x), flags=re.IGNORECASE)[0] if pd.notna(x) else 'Unknown'
                            )
                            live_pivot_data[pivot_req_cols] = live_pivot_data[pivot_req_cols].replace("", "Unknown").fillna("Unknown")
                            
                            live_pivot_df = live_pivot_data.groupby(pivot_req_cols).agg(
                                Device_Count=('POS Name', 'count'),
                                Station_Count=('Station ID', 'nunique')
                            ).reset_index()
                            
                            live_pivot_df.rename(columns={
                                'Device_Count': 'Total POS Devices', 
                                'Station_Count': 'Total Unique Stations'
                            }, inplace=True)
                            
                            live_pivot_df = live_pivot_df.sort_values(by=['POS Manufacture', 'ApplicationROMID (FW)'])
                            st.dataframe(live_pivot_df, width='stretch', hide_index=True)
                        else:
                            st.info("Pivot summary unavailable. Missing required columns.")

                        # --- EXPORT & SAVE ACTIONS ---
                        scanner_excel_data = generate_excel_bytes(clean_scanner_df)
                        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                        scanner_filename = f"Scanner_Recon_{timestamp_str}.xlsx"
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        col_act1, col_act2 = st.columns([3, 1])
                        with col_act1:
                            if st.button("💾 Save Reconciled Scan to Database", key="save_scanner_db", type="primary"):
                                save_run_to_db("Scanner_Recon", scanner_filename, scanner_row_count, scanner_excel_data.getvalue())
                                st.toast("Scanner recon saved to database history.", icon="✅")
                                time.sleep(1)
                                st.session_state.scanner_counter += 1
                                st.rerun()
                        with col_act2:
                            if st.button("📥 Export to Output Folder", key="scanner_dl_btn"):
                                filepath = export_to_local_output(scanner_excel_data, scanner_filename)
                                st.toast(f"Exported directly to {filepath}!", icon="🚀")
                                time.sleep(1)
                                st.session_state.scanner_counter += 1
                                st.rerun()

            except Exception as e:
                st.error(f"❌ Scanner Processing Error: {e}")

    st.markdown("<br>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("🗄️ Scanner Reconciliation History")
        scanner_history_df = get_history_metadata("Scanner_Recon")
        
        if scanner_history_df.empty:
            st.caption("No historical scanner reconciliations found.")
        else:
            for index, row in scanner_history_df.iterrows():
                r_id = row['id']
                r_col1, r_col2, r_col3, r_col4, r_col5, r_col6 = st.columns([2, 3, 1, 1, 1, 1], vertical_alignment="center")
                with r_col1: st.caption(str(row['run_date']))
                with r_col2: st.caption(str(row['filename']))
                with r_col3: st.caption(f"{row['record_count']}")
                
                db_file_data, db_filename = get_file_from_db(r_id)
                with r_col4:
                    if db_file_data and st.button("📥 Export", key=f"dl_scanner_{r_id}", use_container_width=True):
                        export_to_local_output(db_file_data, db_filename)
                        st.toast("Exported to local volume.", icon="🚀")
                        time.sleep(1)
                        st.rerun()
                with r_col5:
                    if st.button("👁️ View", key=f"rev_scanner_{r_id}", use_container_width=True):
                        st.session_state[f"show_review_{r_id}"] = not st.session_state.get(f"show_review_{r_id}", False)
                with r_col6:
                    if st.button("🗑️ Drop", key=f"del_scanner_{r_id}", use_container_width=True):
                        delete_record_from_db(r_id)
                        st.rerun()
                
                if st.session_state.get(f"show_review_{r_id}", False) and db_file_data:
                    st.info(f"Viewing Snapshot: {db_filename}")
                    st.dataframe(pd.read_excel(io.BytesIO(db_file_data)), width='stretch')
                st.divider()